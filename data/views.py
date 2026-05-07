from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Avg, Max
from django.utils import timezone
from django.http import HttpResponse
from math import sqrt
from datetime import timedelta
import json
import math
import openpyxl
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from .models import Supplier, KelolaData, HasilPrediksi, JENIS_BERAS_TETAP

def generate_no_pemasok():
    last_supplier = Supplier.objects.order_by('-id').first()
    next_id = 1 if not last_supplier else last_supplier.id + 1
    return f"SUP-{next_id:03d}"

def landing(request):
    return render(request, 'landing.html')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')

        messages.error(request, 'Username atau password salah')

    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def dashboard(request):
    data = KelolaData.objects.filter(jenis_beras=JENIS_BERAS_TETAP).order_by('tanggal')

    range_hari = request.GET.get('range')

    if range_hari:
        today = timezone.now().date()
        start_date = today - timedelta(days=int(range_hari))
        data = data.filter(tanggal__gte=start_date)

    total_data = data.count()
    total_supplier = Supplier.objects.count()
    harga_tertinggi = data.aggregate(Max('harga_per_kg'))['harga_per_kg__max']

    tanggal = []
    harga = []

    for d in data:
        tanggal.append(d.tanggal.strftime("%Y-%m-%d"))
        harga.append(d.harga_per_kg)

    hasil_terbaru = HasilPrediksi.objects.order_by('-id').first()

    prediksi = harga.copy()

    if hasil_terbaru:
        prediksi.append(round(hasil_terbaru.harga_prediksi))
        prediksi_harga = round(hasil_terbaru.harga_prediksi)
    elif harga:
        prediksi.append(harga[-1] + 200)
        prediksi_harga = prediksi[-1]
    else:
        prediksi_harga = 0

    paginator = Paginator(data.order_by('-tanggal'), 5)
    page_number = request.GET.get('page')
    data_list = paginator.get_page(page_number)

    context = {
        'total_data': total_data,
        'total_supplier': total_supplier,
        'harga_tertinggi': harga_tertinggi,
        'prediksi_harga': prediksi_harga,
        'tanggal': json.dumps(tanggal),
        'harga': json.dumps(harga),
        'prediksi': json.dumps(prediksi),
        'data_list': data_list,
        'jenis_beras_tetap': JENIS_BERAS_TETAP,
    }

    return render(request, 'dashboard/index.html', context)


@login_required(login_url='login')
def supplier(request):
    supplier_list = Supplier.objects.all().order_by('-id')

    paginator = Paginator(supplier_list, 5)
    page_number = request.GET.get('page')
    suppliers = paginator.get_page(page_number)

    return render(request, 'supplier/index.html', {
        'suppliers': suppliers
    })


@login_required(login_url='login')
def supplier_create(request):

    if request.method == 'POST':

        no_pemasok = request.POST.get('no_pemasok')
        nama_supplier = request.POST.get('nama_supplier')
        kontak = request.POST.get('kontak')
        alamat = request.POST.get('alamat')
        catatan = request.POST.get('catatan')

        # cek duplicate
        if Supplier.objects.filter(no_pemasok=no_pemasok).exists():
            messages.error(request, "No pemasok sudah digunakan!")
            return render(request, 'supplier/create.html')

        # simpan data
        Supplier.objects.create(
            no_pemasok=no_pemasok,
            nama_supplier=nama_supplier,
            kontak=kontak,
            alamat=alamat,
            catatan=catatan,
        )

        messages.success(request, "Supplier berhasil ditambahkan!")
        return redirect('supplier')

    return render(request, 'supplier/create.html')


@login_required(login_url='login')
def supplier_detail(request, id):
    supplier = get_object_or_404(Supplier, id=id)
    return render(request, 'supplier/detail.html', {'supplier': supplier})


@login_required(login_url='login')
def supplier_edit(request, id):
    supplier = get_object_or_404(Supplier, id=id)

    if request.method == 'POST':
        supplier.no_pemasok = request.POST.get('no_pemasok')
        supplier.nama_supplier = request.POST.get('nama_supplier')
        supplier.kontak = request.POST.get('kontak')
        supplier.alamat = request.POST.get('alamat')
        supplier.catatan = request.POST.get('catatan')
        supplier.save()
        return redirect('supplier')

    return render(request, 'supplier/edit.html', {'supplier': supplier})


@login_required(login_url='login')
def supplier_delete(request, id):
    supplier = get_object_or_404(Supplier, id=id)
    supplier.delete()
    return redirect('supplier')


@login_required(login_url='login')
def kelola_data(request):
    data_list = KelolaData.objects.filter(
        jenis_beras=JENIS_BERAS_TETAP
    ).order_by('-id')

    paginator = Paginator(data_list, 5)  # 5 data per halaman
    page_number = request.GET.get('page')
    data = paginator.get_page(page_number)

    return render(request, 'kelola_data/index.html', {
        'data': data
    })

@login_required(login_url='login')
def kelola_data_create(request):
    suppliers = Supplier.objects.all()

    if request.method == 'POST':
        KelolaData.objects.create(
            tanggal=request.POST.get('tanggal'),
            jenis_beras=JENIS_BERAS_TETAP,
            supplier=Supplier.objects.get(id=request.POST.get('supplier')),
            kuantitas_kg=request.POST.get('kuantitas_kg'),
            stok_kg=request.POST.get('stok_kg'),
            harga_per_kg=request.POST.get('harga_per_kg'),
        )
        return redirect('kelola_data')

    return render(request, 'kelola_data/create.html', {
        'suppliers': suppliers,
        'jenis_beras_tetap': JENIS_BERAS_TETAP,
    })


@login_required(login_url='login')
def kelola_data_edit(request, id):
    data = get_object_or_404(KelolaData, id=id)
    suppliers = Supplier.objects.all()

    if request.method == 'POST':
        data.tanggal = request.POST.get('tanggal')
        data.jenis_beras = JENIS_BERAS_TETAP
        data.supplier = Supplier.objects.get(id=request.POST.get('supplier'))
        data.kuantitas_kg = request.POST.get('kuantitas_kg')
        data.stok_kg = request.POST.get('stok_kg')
        data.harga_per_kg = request.POST.get('harga_per_kg')
        data.save()
        return redirect('kelola_data')

    return render(request, 'kelola_data/edit.html', {
        'data': data,
        'suppliers': suppliers,
        'jenis_beras_tetap': JENIS_BERAS_TETAP,
    })


@login_required(login_url='login')
def kelola_data_delete(request, id):
    data = get_object_or_404(KelolaData, id=id)
    data.delete()
    return redirect('kelola_data')

@login_required(login_url='login')
def download_template_kelola_data(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Import"

    ws.append([
        "tanggal",
        "jenis_beras",
        "supplier",
        "kuantitas_kg",
        "harga_per_kg",
        "stok_kg"
    ])

    ws.append([
        "2026-01-01",
        JENIS_BERAS_TETAP,
        "Supplier A",
        100,
        12000,
        80
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=template_kelola_data.xlsx'

    wb.save(response)
    return response


@login_required(login_url='login')
def import_kelola_data_excel(request):
    if request.method == "POST":
        excel_file = request.FILES.get("file")

        if not excel_file:
            messages.error(request, "File belum dipilih!")
            return redirect("kelola_data")

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            tanggal = row[0]
            supplier_nama = row[2]
            kuantitas = row[3]
            harga = row[4]
            stok = row[5]

            supplier, created = Supplier.objects.get_or_create(
                nama_supplier=supplier_nama,
                defaults={
                    'no_pemasok': generate_no_pemasok(),
                    'kontak': '-',
                    'alamat': '-',
                    'catatan': 'Dibuat otomatis dari import Excel'
                }
            )

            KelolaData.objects.create(
                tanggal=tanggal,
                jenis_beras=JENIS_BERAS_TETAP,
                supplier=supplier,
                kuantitas_kg=kuantitas,
                harga_per_kg=harga,
                stok_kg=stok
            )

        messages.success(request, "Import data berhasil!")
        return redirect("kelola_data")

    return redirect("kelola_data")


@login_required(login_url='login')
def export_kelola_data_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="laporan_data_beras.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph("Laporan Data Beras", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))

    data_table = []
    data_table.append([
        "No",
        "Tanggal",
        "Jenis Beras",
        "Supplier",
        "Kuantitas",
        "Harga/kg",
        "Stok"
    ])

    data = KelolaData.objects.filter(jenis_beras=JENIS_BERAS_TETAP).order_by('-tanggal')

    for i, d in enumerate(data, 1):
        data_table.append([
            i,
            str(d.tanggal),
            d.jenis_beras,
            d.supplier.nama_supplier,
            f"{d.kuantitas_kg} kg",
            f"Rp {d.harga_per_kg}",
            f"{d.stok_kg} kg"
        ])

    table = Table(data_table)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


@login_required(login_url='login')
def dataset(request):
    data = KelolaData.objects.filter(jenis_beras=JENIS_BERAS_TETAP).order_by('-tanggal')
    return render(request, 'dataset/index.html', {
        'data': data,
        'jenis_beras_tetap': JENIS_BERAS_TETAP,
    })


@login_required(login_url='login')
def prediksi(request):
    data = KelolaData.objects.filter(jenis_beras=JENIS_BERAS_TETAP).order_by('-tanggal')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        data = data.filter(tanggal__range=[start_date, end_date])

    avg_harga = data.aggregate(Avg('harga_per_kg'))['harga_per_kg__avg']
    avg_stok = data.aggregate(Avg('stok_kg'))['stok_kg__avg']

    context = {
        'data': data,
        'total_data': data.count(),
        'avg_harga': round(avg_harga or 0),
        'avg_stok': round(avg_stok or 0),
        'jenis_beras_tetap': JENIS_BERAS_TETAP,
    }

    return render(request, 'prediksi/index.html', context)


@login_required(login_url='login')
def proses_prediksi(request):
    if request.method != 'POST':
        return redirect('prediksi')

    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')

    data_histori = KelolaData.objects.filter(
        jenis_beras=JENIS_BERAS_TETAP
    ).order_by('-tanggal')

    if start_date and end_date:
        data_histori = data_histori.filter(tanggal__range=[start_date, end_date])

    if data_histori.count() < 3:
        messages.error(request, "Data tidak cukup untuk prediksi")
        return redirect('prediksi')

    try:
        stok_target = float(request.POST.get('target_stok'))
        kuantitas_target = float(request.POST.get('target_kuantitas'))
    except (TypeError, ValueError):
        messages.error(request, "Input target stok dan target kuantitas harus diisi dengan benar")
        return redirect('prediksi')

    tanggal = timezone.now().date()
    jenis_beras = JENIS_BERAS_TETAP
    total_data = data_histori.count()

    k = 3

    perhitungan_detail = []

    for data in data_histori:
        selisih_stok = data.stok_kg - stok_target
        selisih_kuantitas = data.kuantitas_kg - kuantitas_target

        kuadrat_stok = selisih_stok ** 2
        kuadrat_kuantitas = selisih_kuantitas ** 2
        jumlah_kuadrat = kuadrat_stok + kuadrat_kuantitas
        jarak = sqrt(jumlah_kuadrat)

        perhitungan_detail.append({
            "stok": data.stok_kg,
            "kuantitas": data.kuantitas_kg,
            "harga": data.harga_per_kg,
            "jarak": jarak,
            "selisih_stok": selisih_stok,
            "selisih_kuantitas": selisih_kuantitas,
            "kuadrat_stok": kuadrat_stok,
            "kuadrat_kuantitas": kuadrat_kuantitas,
            "jumlah_kuadrat": jumlah_kuadrat,
        })

    perhitungan_detail = sorted(perhitungan_detail, key=lambda x: x['jarak'])
    tetangga = perhitungan_detail[:k]

    total_harga = sum(t['harga'] for t in tetangga)
    hasil = total_harga / k

    total_error = 0
    rmse_detail = []

    for t in tetangga:
        error = t['harga'] - hasil
        kuadrat_error = error ** 2
        total_error += kuadrat_error

        rmse_detail.append({
            "harga_asli": t['harga'],
            "error": error,
            "kuadrat_error": kuadrat_error
        })

    mse = total_error / k
    rmse = sqrt(mse)

    detail_json = json.dumps({
        "target": {
            "stok": stok_target,
            "kuantitas": kuantitas_target
        },
        "perhitungan": perhitungan_detail,
        "tetangga": tetangga,
        "rmse_detail": rmse_detail,
        "total_harga": total_harga,
        "mse": mse
    })


    hasil, created = HasilPrediksi.objects.update_or_create(
    tanggal=tanggal,
    defaults={
        'jenis_beras': JENIS_BERAS_TETAP,
        'stok_target': stok_target,
        'kuantitas_target': kuantitas_target,
        'harga_prediksi': hasil,
        'nilai_rmse': rmse,
        'k_digunakan': k,
        'detail_perhitungan': detail_json,
    }
)
    return redirect('hasil_prediksi')

@login_required(login_url='login')
def hasil(request):

    hasil_list = HasilPrediksi.objects.all().order_by('-id')

    paginator = Paginator(hasil_list, 5) 
    page_number = request.GET.get('page')

    data = paginator.get_page(page_number)

    return render(request, 'hasil/index.html', {
        'data': data
    })



@login_required(login_url='login')
def hasil_prediksi(request):
    data = HasilPrediksi.objects.order_by('-id').first()

    if not data:
        return render(request, 'prediksi/hasil_list.html', {
            'data': None,
            'detail': None
        })

    import json
    detail = json.loads(data.detail_perhitungan)

    return render(request, 'prediksi/hasil_list.html', {
        'data': data,
        'detail': detail
    })

